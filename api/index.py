from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import os
import json
import tempfile
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import httpx
import base64

app = FastAPI()

# CORS 配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 临时存储用户数据（仅用于演示，实际应使用数据库）
users_db = {}
templates_db = {}
generation_records = []

# ============ 用户相关 API ============

class UserRequest(BaseModel):
    phone: str
    password: str

@app.post("/api/users/register")
async def register(req: UserRequest):
    """用户注册"""
    if req.phone in users_db:
        return {"code": 1, "message": "手机号已注册"}
    
    users_db[req.phone] = {"password": req.password, "phone": req.phone}
    return {
        "code": 0,
        "data": {"token": f"token_{req.phone}", "phone": req.phone}
    }

@app.post("/api/users/login")
async def login(req: UserRequest):
    """用户登录"""
    if req.phone not in users_db or users_db[req.phone]["password"] != req.password:
        return {"code": 1, "message": "手机号或密码错误"}
    
    return {
        "code": 0,
        "data": {"token": f"token_{req.phone}", "phone": req.phone}
    }

# ============ 模板相关 API ============

@app.post("/api/templates/upload")
async def upload_template(file: UploadFile = File(...)):
    """上传并解析 Excel 模板"""
    try:
        # 保存上传的文件到临时目录
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, file.filename)
        
        # 读取文件内容
        content = await file.read()
        with open(file_path, 'wb') as f:
            f.write(content)
        
        # 使用 openpyxl 解析
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 获取数据范围
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 提取前 10 行作为预览
        preview_data = []
        for row in ws.iter_rows(max_row=min(10, max_row), values_only=False):
            row_data = []
            for cell in row[:max_col]:
                value = cell.value
                row_data.append(str(value) if value is not None else "")
            preview_data.append(row_data)
        
        # 获取合并单元格
        merged_cells = list(ws.merged_cells.ranges)
        merged_cells = [str(mc) for mc in merged_cells]
        
        file_id = f"file_{len(templates_db)}"
        templates_db[file_id] = {
            "file_path": file_path,
            "sheet_name": ws.title,
            "max_row": max_row,
            "max_col": max_col,
            "merged_cells": merged_cells
        }
        
        return {
            "code": 0,
            "data": {
                "file_id": file_id,
                "sheet_name": ws.title,
                "max_row": max_row,
                "max_col": max_col,
                "merged_cells": merged_cells,
                "preview_data": preview_data
            }
        }
    except Exception as e:
        return {
            "code": 1,
            "message": f"文件解析失败: {str(e)}"
        }

class TemplateConfig(BaseModel):
    template_id: str
    template_name: str
    config: dict

@app.post("/api/templates/config")
async def save_template_config(req: TemplateConfig):
    """保存模板配置"""
    try:
        config_id = f"config_{len(generation_records)}"
        templates_db[req.template_id]["config"] = req.config
        templates_db[req.template_id]["name"] = req.template_name
        
        return {
            "code": 0,
            "data": {"config_id": config_id}
        }
    except Exception as e:
        return {
            "code": 1,
            "message": str(e)
        }

@app.get("/api/templates/list")
async def get_templates():
    """获取模板列表"""
    templates = []
    for file_id, data in templates_db.items():
        templates.append({
            "id": file_id,
            "name": data.get("name", "未命名"),
            "sheet_name": data.get("sheet_name", "Sheet1"),
            "created_at": "2024-01-01"
        })
    
    return {
        "code": 0,
        "data": templates
    }

# ============ 合同识别 API ============

@app.post("/api/contracts/extract")
async def extract_contract(file: UploadFile = File(...)):
    """AI 识别合同内容"""
    try:
        content = await file.read()
        
        # 这里调用 AI 大模型（示例使用通义千问）
        # 实际部署时需要配置 API Key
        extracted_data = await call_ai_extract(content, file.filename)
        
        return {
            "code": 0,
            "data": extracted_data
        }
    except Exception as e:
        return {
            "code": 1,
            "message": f"识别失败: {str(e)}"
        }

async def call_ai_extract(file_content: bytes, filename: str) -> dict:
    """调用 AI 模型识别合同"""
    # 这是一个示例实现
    # 实际需要集成真实的 AI API（如通义千问、ChatGPT 等）
    
    # 示例返回数据
    return {
        "buyer": "SALASAR IMPEX",
        "contract_no": "SC260269-SIM",
        "date": "2026-03-02",
        "currency": "USD",
        "items": [
            {"code": "702984", "name": "Track Link Assy", "qty": 4, "unit_price": 630.00},
            {"code": "702507", "name": "Track Link Assy", "qty": 6, "unit_price": 630.00},
        ],
        "total_amount": 5400.00
    }

# ============ 发票生成 API ============

class GenerateRequest(BaseModel):
    config_id: str
    extracted_data: dict
    exchange_rate: float = 7.25

@app.post("/api/invoices/generate")
async def generate_invoice(req: GenerateRequest):
    """生成 PI 发票文件"""
    try:
        # 从存储的模板获取文件路径
        template_data = None
        for file_id, data in templates_db.items():
            if data.get("config") is not None:
                template_data = data
                break
        
        if not template_data:
            return {
                "code": 1,
                "message": "未找到模板"
            }
        
        # 加载原始模板
        template_path = template_data["file_path"]
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        config = template_data.get("config", {})
        header_end_row = config.get("header_end_row", 5)
        footer_start_row = config.get("footer_start_row", 25)
        
        # 填充数据
        items = req.extracted_data.get("items", [])
        
        # 计算需要的行数
        current_detail_rows = footer_start_row - header_end_row - 1
        needed_detail_rows = len(items)
        delta = needed_detail_rows - current_detail_rows
        
        # 动态插入或删除行
        if delta > 0:
            ws.insert_rows(header_end_row + 1, delta)
        elif delta < 0:
            for _ in range(abs(delta)):
                ws.delete_rows(header_end_row + 1)
        
        # 填充字段数据
        field_mapping = config.get("fields", {})
        if "buyer" in field_mapping:
            coord = field_mapping["buyer"]
            ws[coord] = req.extracted_data.get("buyer", "")
        if "invoice_no" in field_mapping:
            coord = field_mapping["invoice_no"]
            ws[coord] = req.extracted_data.get("contract_no", "")
        if "date" in field_mapping:
            coord = field_mapping["date"]
            ws[coord] = req.extracted_data.get("date", "")
        
        # 填充产品明细行
        detail_cols = config.get("detail_columns", {
            "code": "B",
            "name": "C",
            "qty": "D",
            "price": "E"
        })
        
        for idx, item in enumerate(items):
            row_num = header_end_row + 1 + idx
            ws[f"{detail_cols['code']}{row_num}"] = item.get("code", "")
            ws[f"{detail_cols['name']}{row_num}"] = item.get("name", "")
            ws[f"{detail_cols['qty']}{row_num}"] = item.get("qty", 0)
            ws[f"{detail_cols['price']}{row_num}"] = item.get("unit_price", 0)
        
        # 保存到临时文件
        output_path = os.path.join(tempfile.gettempdir(), f"PI_{req.extracted_data.get('buyer', 'export')}.xlsx")
        wb.save(output_path)
        
        # 记录生成历史
        generation_records.append({
            "id": len(generation_records),
            "buyer_name": req.extracted_data.get("buyer", ""),
            "product_count": len(items),
            "exchange_rate": req.exchange_rate,
            "output_file": output_path,
            "created_at": "2024-01-01"
        })
        
        # 返回文件流
        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"PI_{req.extracted_data.get('buyer', 'export')}.xlsx"
        )
    except Exception as e:
        return {
            "code": 1,
            "message": f"生成失败: {str(e)}"
        }

# ============ 历史记录 API ============

@app.get("/api/history/list")
async def get_history(page: int = 1, size: int = 20):
    """获取生成历史记录"""
    start = (page - 1) * size
    end = start + size
    
    records = generation_records[start:end]
    
    return {
        "code": 0,
        "data": records
    }

# ============ 健康检查 ============

@app.get("/api/health")
async def health_check():
    """健康检查"""
    return {"status": "ok"}

# Vercel 部署时需要的导出
from mangum import Mangum
handler = Mangum(app)
